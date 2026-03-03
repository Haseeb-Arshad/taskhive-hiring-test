"""Preview API — serve and render agent workspace files for user inspection.

Supports: code files (with syntax info), markdown, HTML, images, PDFs,
spreadsheets (xlsx → JSON table), CSV, and directory listings.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import mimetypes
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from sqlalchemy import select

from app.config import settings
from app.db.engine import async_session
from app.db.models import OrchSubtask, OrchTaskExecution

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/orchestrator/preview", tags=["preview"])

# Map extensions to preview categories
PREVIEW_CATEGORIES: dict[str, str] = {
    # Code
    ".py": "code", ".js": "code", ".jsx": "code", ".ts": "code", ".tsx": "code",
    ".java": "code", ".go": "code", ".rs": "code", ".rb": "code", ".php": "code",
    ".c": "code", ".cpp": "code", ".h": "code", ".hpp": "code", ".cs": "code",
    ".sh": "code", ".bash": "code", ".zsh": "code", ".fish": "code",
    ".sql": "code", ".r": "code", ".swift": "code", ".kt": "code",
    ".yaml": "code", ".yml": "code", ".toml": "code", ".ini": "code",
    ".cfg": "code", ".conf": "code", ".env": "code",
    ".dockerfile": "code", ".makefile": "code",
    # Markup / documents
    ".md": "markdown", ".markdown": "markdown", ".rst": "markdown",
    ".html": "html", ".htm": "html",
    ".json": "json", ".jsonl": "json",
    ".xml": "code", ".svg": "image",
    ".txt": "text", ".log": "text", ".csv": "csv",
    # Spreadsheets
    ".xlsx": "spreadsheet", ".xls": "spreadsheet",
    # Images
    ".png": "image", ".jpg": "image", ".jpeg": "image", ".gif": "image",
    ".webp": "image", ".bmp": "image", ".ico": "image",
    # PDF
    ".pdf": "pdf",
    # Visualization
    ".ipynb": "notebook",
}

# Language hints for syntax highlighting
LANG_MAP: dict[str, str] = {
    ".py": "python", ".js": "javascript", ".jsx": "jsx", ".ts": "typescript",
    ".tsx": "tsx", ".java": "java", ".go": "go", ".rs": "rust", ".rb": "ruby",
    ".php": "php", ".c": "c", ".cpp": "cpp", ".h": "c", ".hpp": "cpp",
    ".cs": "csharp", ".sh": "bash", ".bash": "bash", ".sql": "sql",
    ".r": "r", ".swift": "swift", ".kt": "kotlin", ".yaml": "yaml",
    ".yml": "yaml", ".toml": "toml", ".json": "json", ".xml": "xml",
    ".html": "html", ".htm": "html", ".css": "css", ".scss": "scss",
    ".dockerfile": "dockerfile", ".makefile": "makefile",
    ".md": "markdown", ".rst": "restructuredtext",
}


def _get_workspace_path(execution: OrchTaskExecution) -> Path:
    """Get the workspace path for an execution."""
    if execution.workspace_path:
        return Path(execution.workspace_path)
    return Path(settings.WORKSPACE_ROOT) / f"task-{execution.id}"


def _safe_resolve(workspace: Path, relative: str) -> Path:
    """Resolve a relative path within the workspace, preventing traversal."""
    target = (workspace / relative).resolve()
    ws_resolved = workspace.resolve()
    if not str(target).startswith(str(ws_resolved)):
        raise ValueError("Path traversal detected")
    return target


def _scan_directory(directory: Path, workspace_root: Path, max_depth: int = 5) -> list[dict]:
    """Recursively scan a directory and return a file tree."""
    entries: list[dict] = []

    if not directory.exists():
        return entries

    try:
        items = sorted(directory.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
    except PermissionError:
        return entries

    for item in items:
        if item.name.startswith(".") and item.name not in (".env.example",):
            continue
        if item.name in ("node_modules", "__pycache__", ".git", ".venv", "venv"):
            continue

        relative = str(item.relative_to(workspace_root)).replace("\\", "/")
        ext = item.suffix.lower()

        if item.is_dir():
            children = []
            if max_depth > 0:
                children = _scan_directory(item, workspace_root, max_depth - 1)
            entries.append({
                "name": item.name,
                "path": relative,
                "type": "directory",
                "children": children,
            })
        elif item.is_file():
            try:
                size = item.stat().st_size
            except OSError:
                size = 0
            entries.append({
                "name": item.name,
                "path": relative,
                "type": "file",
                "size": size,
                "category": PREVIEW_CATEGORIES.get(ext, "binary"),
                "language": LANG_MAP.get(ext, ""),
            })

    return entries


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/executions")
async def list_preview_executions(
    limit: int = 20,
    offset: int = 0,
) -> dict[str, Any]:
    """List executions with their workspace file trees for preview."""
    async with async_session() as session:
        query = (
            select(OrchTaskExecution)
            .order_by(OrchTaskExecution.id.desc())
            .offset(offset)
            .limit(limit)
        )
        result = await session.execute(query)
        executions = result.scalars().all()

    data = []
    for ex in executions:
        ws = _get_workspace_path(ex)
        has_workspace = ws.exists()
        file_count = 0
        if has_workspace:
            try:
                file_count = sum(1 for _ in ws.rglob("*") if _.is_file())
            except Exception:
                pass

        data.append({
            "id": ex.id,
            "taskhive_task_id": ex.taskhive_task_id,
            "status": ex.status,
            "task_title": (ex.task_snapshot or {}).get("title", "Untitled"),
            "task_description": (ex.task_snapshot or {}).get("description", "")[:200],
            "total_tokens_used": ex.total_tokens_used,
            "attempt_count": ex.attempt_count,
            "has_workspace": has_workspace,
            "file_count": file_count,
            "error_message": ex.error_message,
            "created_at": ex.created_at.isoformat() if ex.created_at else None,
            "completed_at": ex.completed_at.isoformat() if ex.completed_at else None,
        })

    return {"ok": True, "data": data}


@router.get("/executions/{execution_id}")
async def get_execution_preview(execution_id: int) -> dict[str, Any]:
    """Full execution details with file tree and subtask results."""
    async with async_session() as session:
        result = await session.execute(
            select(OrchTaskExecution).where(OrchTaskExecution.id == execution_id)
        )
        execution = result.scalar_one_or_none()
        if not execution:
            raise HTTPException(404, "Execution not found")

        # Get subtasks
        sub_result = await session.execute(
            select(OrchSubtask)
            .where(OrchSubtask.execution_id == execution_id)
            .order_by(OrchSubtask.order_index)
        )
        subtasks = sub_result.scalars().all()

    ws = _get_workspace_path(execution)
    file_tree = _scan_directory(ws, ws) if ws.exists() else []

    return {
        "ok": True,
        "data": {
            "id": execution.id,
            "taskhive_task_id": execution.taskhive_task_id,
            "status": execution.status,
            "task_snapshot": execution.task_snapshot,
            "workspace_path": str(ws),
            "total_tokens_used": execution.total_tokens_used,
            "total_cost_usd": execution.total_cost_usd,
            "attempt_count": execution.attempt_count,
            "error_message": execution.error_message,
            "claimed_credits": execution.claimed_credits,
            "file_tree": file_tree,
            "subtasks": [
                {
                    "id": st.id,
                    "order_index": st.order_index,
                    "title": st.title,
                    "description": st.description,
                    "status": st.status,
                    "result": st.result,
                    "files_changed": st.files_changed,
                }
                for st in subtasks
            ],
            "created_at": execution.created_at.isoformat() if execution.created_at else None,
            "completed_at": execution.completed_at.isoformat() if execution.completed_at else None,
        },
    }


@router.get("/executions/{execution_id}/files")
async def list_execution_files(
    execution_id: int,
    path: str = "",
) -> dict[str, Any]:
    """List files in an execution's workspace directory."""
    async with async_session() as session:
        result = await session.execute(
            select(OrchTaskExecution).where(OrchTaskExecution.id == execution_id)
        )
        execution = result.scalar_one_or_none()
        if not execution:
            raise HTTPException(404, "Execution not found")

    ws = _get_workspace_path(execution)
    if not ws.exists():
        return {"ok": True, "data": [], "message": "Workspace not found or cleaned up"}

    try:
        target = _safe_resolve(ws, path) if path else ws
    except ValueError:
        raise HTTPException(400, "Invalid path")

    if not target.exists():
        raise HTTPException(404, "Path not found in workspace")

    if target.is_file():
        ext = target.suffix.lower()
        return {
            "ok": True,
            "data": [{
                "name": target.name,
                "path": str(target.relative_to(ws)).replace("\\", "/"),
                "type": "file",
                "size": target.stat().st_size,
                "category": PREVIEW_CATEGORIES.get(ext, "binary"),
                "language": LANG_MAP.get(ext, ""),
            }],
        }

    tree = _scan_directory(target, ws, max_depth=1)
    return {"ok": True, "data": tree}


@router.get("/executions/{execution_id}/file")
async def get_file_content(
    execution_id: int,
    path: str = Query(..., description="Relative file path within workspace"),
) -> Any:
    """Get the content of a specific file for preview.

    Returns JSON with content + metadata for text files,
    or raw file for binary (images, PDFs).
    """
    async with async_session() as session:
        result = await session.execute(
            select(OrchTaskExecution).where(OrchTaskExecution.id == execution_id)
        )
        execution = result.scalar_one_or_none()
        if not execution:
            raise HTTPException(404, "Execution not found")

    ws = _get_workspace_path(execution)
    if not ws.exists():
        raise HTTPException(404, "Workspace not found or cleaned up")

    try:
        target = _safe_resolve(ws, path)
    except ValueError:
        raise HTTPException(400, "Invalid path")

    if not target.exists():
        raise HTTPException(404, f"File not found: {path}")

    if not target.is_file():
        raise HTTPException(400, "Path is a directory, not a file")

    ext = target.suffix.lower()
    category = PREVIEW_CATEGORIES.get(ext, "binary")

    # Binary files: serve raw
    if category in ("image", "pdf", "binary"):
        mime = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        return FileResponse(str(target), media_type=mime, filename=target.name)

    # Spreadsheets: convert to JSON table
    if category == "spreadsheet":
        return _render_spreadsheet(target)

    # CSV: parse to JSON table
    if category == "csv":
        return _render_csv(target)

    # Notebook: return JSON structure
    if category == "notebook":
        return _render_notebook(target)

    # Text-based files: return content with metadata
    size = target.stat().st_size
    if size > 2_000_000:  # 2MB limit for text preview
        return JSONResponse({
            "ok": True,
            "data": {
                "path": path,
                "category": category,
                "language": LANG_MAP.get(ext, ""),
                "size": size,
                "truncated": True,
                "content": f"File too large for preview ({size:,} bytes). Download instead.",
            },
        })

    try:
        raw = target.read_bytes()
        if b"\x00" in raw[:8192]:
            mime = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
            return FileResponse(str(target), media_type=mime, filename=target.name)
        content = raw.decode("utf-8", errors="replace")
    except Exception as exc:
        raise HTTPException(500, f"Cannot read file: {exc}")

    return JSONResponse({
        "ok": True,
        "data": {
            "path": path,
            "name": target.name,
            "category": category,
            "language": LANG_MAP.get(ext, ""),
            "size": size,
            "line_count": content.count("\n") + 1,
            "content": content,
        },
    })


@router.get("/executions/{execution_id}/download")
async def download_file(
    execution_id: int,
    path: str = Query(..., description="Relative file path within workspace"),
) -> FileResponse:
    """Download a raw file from the workspace."""
    async with async_session() as session:
        result = await session.execute(
            select(OrchTaskExecution).where(OrchTaskExecution.id == execution_id)
        )
        execution = result.scalar_one_or_none()
        if not execution:
            raise HTTPException(404, "Execution not found")

    ws = _get_workspace_path(execution)
    try:
        target = _safe_resolve(ws, path)
    except ValueError:
        raise HTTPException(400, "Invalid path")

    if not target.exists() or not target.is_file():
        raise HTTPException(404, "File not found")

    mime = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
    return FileResponse(str(target), media_type=mime, filename=target.name)


# ---------------------------------------------------------------------------
# Renderers for special file types
# ---------------------------------------------------------------------------

def _render_spreadsheet(file_path: Path) -> JSONResponse:
    """Convert xlsx to a JSON table for preview."""
    try:
        import openpyxl
    except ImportError:
        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "spreadsheet",
                "error": "openpyxl not installed. Install with: pip install openpyxl",
            },
        })

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        sheets: dict[str, Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            # First row as headers
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            sheets[sheet_name] = {
                "headers": headers,
                "rows": data_rows[:500],  # Limit rows for preview
                "total_rows": len(data_rows),
                "truncated": len(data_rows) > 500,
            }
        wb.close()

        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "spreadsheet",
                "sheets": sheets,
            },
        })
    except Exception as exc:
        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "spreadsheet",
                "error": f"Failed to parse spreadsheet: {exc}",
            },
        })


def _render_csv(file_path: Path) -> JSONResponse:
    """Parse CSV to JSON table."""
    try:
        content = file_path.read_text(encoding="utf-8", errors="replace")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "csv",
                "headers": headers,
                "rows": data_rows[:1000],
                "total_rows": len(data_rows),
                "truncated": len(data_rows) > 1000,
            },
        })
    except Exception as exc:
        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "csv",
                "error": f"Failed to parse CSV: {exc}",
            },
        })


def _render_notebook(file_path: Path) -> JSONResponse:
    """Parse Jupyter notebook to structured JSON."""
    try:
        nb = json.loads(file_path.read_text(encoding="utf-8"))
        cells = []
        for cell in nb.get("cells", []):
            cells.append({
                "cell_type": cell.get("cell_type", "code"),
                "source": "".join(cell.get("source", [])),
                "outputs": [
                    _simplify_output(out) for out in cell.get("outputs", [])
                ],
            })

        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "notebook",
                "kernel": nb.get("metadata", {}).get("kernelspec", {}).get("display_name", ""),
                "cells": cells,
            },
        })
    except Exception as exc:
        return JSONResponse({
            "ok": True,
            "data": {
                "path": str(file_path.name),
                "category": "notebook",
                "error": f"Failed to parse notebook: {exc}",
            },
        })


def _simplify_output(output: dict) -> dict:
    """Simplify a notebook cell output for preview."""
    out_type = output.get("output_type", "")
    if out_type == "stream":
        return {"type": "text", "text": "".join(output.get("text", []))}
    elif out_type in ("display_data", "execute_result"):
        data = output.get("data", {})
        if "image/png" in data:
            return {"type": "image", "format": "png", "data": data["image/png"]}
        if "text/html" in data:
            return {"type": "html", "html": "".join(data["text/html"])}
        if "text/plain" in data:
            return {"type": "text", "text": "".join(data["text/plain"])}
    elif out_type == "error":
        return {
            "type": "error",
            "ename": output.get("ename", ""),
            "evalue": output.get("evalue", ""),
            "traceback": output.get("traceback", []),
        }
    return {"type": "unknown"}
